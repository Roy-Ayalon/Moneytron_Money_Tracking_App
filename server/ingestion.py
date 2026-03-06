# server/ingestion.py
"""
File parsing engine: XLS / XLSX / CSV  →  normalized transaction rows.
Ported from client/index.html (extractFromWorkbook_v6 & helpers).
"""

import io
import re
import csv
import math
import logging
from datetime import date

from utils import (
    clean_html, norm_az, norm_heb, numify, parse_date_flex,
    is_text, _is_likely_excel_serial, _excel_serial_to_date, parse_amount,
)

logger = logging.getLogger("moneytron.ingestion")

# ---------------------------------------------------------------------------
# Try to import parsing libraries — graceful fallback messages
# ---------------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None


# ---------------------------------------------------------------------------
# Balance / summary-row detection
# ---------------------------------------------------------------------------
_BALANCE_RE = re.compile(r"(יתרה|balance|saldo|יתרתחשבון|יתרהפתיחה|יתרהסופית)")
_TOTAL_RE = re.compile(r'(סה"?כ|סהכ|סיכום|מאזן|יתרה\s*(פתיחה|סופית)?)')


def _is_balance_label(lbl):
    n = norm_heb(lbl or "")
    return bool(_BALANCE_RE.search(n))


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def _map_from_header_row(matrix, r):
    """
    Given matrix and a header-row index *r*, return a dict with column indices
    for: dateCol, nameCol, transCol, debitCol, hCol, zCol, hzCol.
    """
    row = matrix[r] if r < len(matrix) else []
    norm = [norm_az(c) for c in row]

    def idx(pred):
        for i, n in enumerate(norm):
            if pred(n, i):
                return i
        return -1

    # DATE column
    date_col = idx(lambda x, _: (
        "תאריךהעסקה" in x or x == "תאריך"
        or "תפעולה" in x or "תערך" in x
    ))
    if date_col < 0:
        date_col = idx(lambda x, _: "תאריך" in x or "date" in x)

    # NAME column
    name_col = idx(lambda x, _: (
        ("תיאורהתנועה" in x or "תיאורתנועה" in x)
        and "ערוץ" not in x and "עמלה" not in x and "אסמכתא" not in x
    ))
    if name_col < 0:
        name_col = idx(lambda x, _: (
            "ערוץ" not in x and "ביצוע" not in x
            and "עמלה" not in x and "אסמכתא" not in x and "יתרה" not in x
            and (
                "שםביתעסק" in x or "שםעסק" in x or ("שם" in x and "עסק" in x)
                or "תיאור" in x or "פרטים" in x or "תיאורפעולה" in x
                or "פרטיתנועה" in x or "פירוט" in x
                or "description" in x or "תנועה" in x
            )
        ))

    h_col = idx(lambda x, _: x == "חובה")
    z_col = idx(lambda x, _: x == "זכות")
    hz_col = idx(lambda x, _: "זכותחובה" in x)

    debit_col = idx(lambda x, _: (
        ("סכוםחיוב" in x or "לתשלום" in x or "לחיוב" in x or "בשח" in x)
        and not _is_balance_label(x)
    ))
    trans_col = idx(lambda x, _: (
        ("סכוםעסקה" in x or "amount" in x)
        and not _is_balance_label(x)
    ))

    # Fallback: detect name column by scanning data rows
    if name_col < 0:
        cols = len(row)
        best_c, best_score = -1, -1
        for c in range(cols):
            if c == date_col:
                continue
            text_hits = 0
            date_hits = 0
            len_sum = 0
            for rr in range(r + 1, min(len(matrix), r + 200)):
                v = matrix[rr][c] if c < len(matrix[rr]) else None
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                if parse_date_flex(v):
                    date_hits += 1
                    continue
                if re.search(r"[A-Za-z\u0590-\u05FF]", s):
                    cleaned = re.sub(r"[,\s]", "", s)
                    try:
                        float(cleaned)
                    except ValueError:
                        text_hits += 1
                        len_sum += len(s)
            score = text_hits * 2 + len_sum * 0.05 - date_hits * 4
            if score > best_score:
                best_c, best_score = c, score
        name_col = best_c

    return {
        "row": r,
        "dateCol": date_col if date_col >= 0 else None,
        "nameCol": name_col if name_col >= 0 else None,
        "transCol": trans_col if trans_col >= 0 else None,
        "debitCol": debit_col if debit_col >= 0 else None,
        "hCol": h_col if h_col >= 0 else None,
        "zCol": z_col if z_col >= 0 else None,
        "hzCol": hz_col if hz_col >= 0 else None,
    }


def _find_header_rows(matrix):
    """Return list of row indices that look like header rows."""
    rows = []
    for r, row in enumerate(matrix):
        norm = [norm_az(c) for c in row]
        has_date = any("תאריך" in x or "date" in x or "תפעולה" in x or "תערך" in x for x in norm)
        has_name = any(
            "תיאורהתנועה" in x or "תיאורתנועה" in x
            or "שםביתעסק" in x or "שםעסק" in x or ("שם" in x and "עסק" in x)
            or "תיאור" in x or "פרטים" in x or "תיאורפעולה" in x
            or "פרטיתנועה" in x or "פירוט" in x or "תנועה" in x
            for x in norm
        )
        has_txn = any(
            "סכוםהעסקה" in x or "סכוםעסקה" in x or ("עסקה" in x and "סכום" in x)
            for x in norm
        )
        has_deb = any(
            "סכוםהחיוב" in x or "סכוםחיוב" in x
            or "לתשלום" in x or "לחיוב" in x or "בשח" in x
            for x in norm
        )
        has_hz = any(x == "חובה" or x == "זכות" or "זכותחובה" in x for x in norm)
        if has_date and (has_name or has_txn or has_deb or has_hz):
            rows.append(r)
    return rows


def _detect_heuristic(matrix):
    """Fallback column detection for headerless tables."""
    rows = len(matrix)
    cols = 0
    for row in matrix:
        if len(row) > cols:
            cols = len(row)
    if not cols:
        return None

    date_score = [0] * cols
    text_score = [0] * cols
    for row in matrix:
        for c in range(min(len(row), cols)):
            v = row[c]
            if parse_date_flex(v):
                date_score[c] += 1
            if is_text(v):
                text_score[c] += 1

    date_col = max(range(cols), key=lambda c: date_score[c]) if cols else 0
    name_col = -1
    max_t = -1
    for c in range(cols):
        if c == date_col:
            continue
        if text_score[c] > max_t:
            max_t = text_score[c]
            name_col = c

    # Find numeric "amount" columns
    tx_rows = []
    for ri, row in enumerate(matrix):
        d = parse_date_flex(row[date_col] if date_col < len(row) else None)
        n = row[name_col] if 0 <= name_col < len(row) else None
        nt = str(n or "").strip()
        if d and nt:
            try:
                float(nt)
            except ValueError:
                tx_rows.append(ri)

    cand = []
    for c in range(cols):
        if c == date_col or c == name_col:
            continue
        num = 0
        dec = 0
        for ri in tx_rows:
            row = matrix[ri]
            v = row[c] if c < len(row) else None
            if v is None or str(v).strip() == "":
                continue
            n = numify(v)
            if n is None:
                continue
            num += 1
            if round(n) != n:
                dec += 1
        if not num:
            continue
        cand.append({"ci": c, "score": num + (dec / num) * 2})

    cand.sort(key=lambda x: (-x["score"], -x["ci"]))
    left = min(cand[0]["ci"], (cand[1] if len(cand) > 1 else cand[0])["ci"]) if cand else None
    right = max(cand[0]["ci"], (cand[1] if len(cand) > 1 else cand[0])["ci"]) if cand else None

    return {
        "dateCol": date_col,
        "nameCol": name_col if name_col >= 0 else None,
        "transCol": left,
        "debitCol": right,
    }


# ---------------------------------------------------------------------------
# Row helpers
# ---------------------------------------------------------------------------

def _best_name_from_row(row, avoid_cols=None):
    """Pick the best merchant-name cell from a row, skipping columns in avoid_cols."""
    avoid = set(c for c in (avoid_cols or []) if c is not None)
    best_text, best_score = "", -1
    for c, raw in enumerate(row):
        if c in avoid:
            continue
        if raw is None:
            continue
        s = clean_html(raw)
        if not s:
            continue
        if re.match(r"^[:.\-_,;!?]+$", s):
            continue
        # Skip if it looks like a date
        if parse_date_flex(raw):
            continue
        if re.search(r"\d{4}", s) or re.search(r"GMT|UTC", s):
            continue
        if re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|Mon|Tue|Wed|Thu|Fri|Sat|Sun)", s, re.IGNORECASE):
            continue
        # Skip if mostly numbers
        if re.match(r"^[\d,.\s\-₪$€£]+$", s):
            continue
        # Skip garbage
        if re.match(r"^[x\s«»§=íï\-]+$", s, re.IGNORECASE):
            continue
        # Must have letters
        if not re.search(r"[A-Za-z\u0590-\u05FF]", s):
            continue
        cleaned = re.sub(r"[,\s₪$€£]", "", s)
        try:
            float(cleaned)
            if not re.search(r"[A-Za-z\u0590-\u05FF]", cleaned):
                continue
        except ValueError:
            pass
        letter_count = len(re.findall(r"[A-Za-z\u0590-\u05FF]", s))
        score = letter_count * 10 + len(s)
        if score > best_score:
            best_text, best_score = s, score
    return best_text


def _is_header_row(row):
    """Quick detector for a header row."""
    n = [norm_az(c) for c in row]
    has_date = any("תאריך" in x or "date" in x or "תפעולה" in x or "תערך" in x for x in n)
    has_name = any(
        "תיאורהתנועה" in x or "תיאורתנועה" in x
        or "שםביתעסק" in x or "שםעסק" in x or ("שם" in x and "עסק" in x)
        or "תיאור" in x or "פרטים" in x or "תיאורפעולה" in x
        or "פרטיתנועה" in x or "פירוט" in x or "description" in x
        or "תנועה" in x or "אסמכתא" in x
        for x in n
    )
    has_txn = any("סכוםעסקה" in x or ("עסקה" in x and "סכום" in x) or "amount" in x for x in n)
    has_deb = any("סכוםחיוב" in x or "לתשלום" in x or "לחיוב" in x or "בשח" in x or "סכום" in x for x in n)
    has_hz = any(x == "חובה" or x == "זכות" or "זכותחובה" in x for x in n)
    return has_date and (has_name or has_txn or has_deb or has_hz)


def _to_iso(d):
    """Convert a parsed date to ISO string."""
    if isinstance(d, date):
        return d.strftime("%Y-%m-%d")
    if isinstance(d, str):
        dd = parse_date_flex(d)
        if dd:
            return dd.strftime("%Y-%m-%d")
    if isinstance(d, (int, float)) and _is_likely_excel_serial(d):
        dd = _excel_serial_to_date(d)
        if dd:
            return dd.strftime("%Y-%m-%d")
    return None


# ---------------------------------------------------------------------------
# Sheet extractor
# ---------------------------------------------------------------------------

def _extract_from_sheet(matrix, sheet_name, file_name, user_tag=None):
    """
    Parse a 2D matrix (list of lists) into normalised transaction dicts.
    user_tag: optional month tag (1-12) for date validation/swap.
    """
    rows_out = []

    def push_row(date_obj, name, t, db, credit_by_hz):
        iso = _to_iso(date_obj)
        name_str = str(name or "").strip()
        t_num = float(t) if t is not None else None
        db_num = float(db) if db is not None else None
        base = db_num if (db_num is not None and db_num != 0) else (t_num if t_num is not None else 0)
        if not iso or not name_str or not base:
            return

        credit = credit_by_hz if credit_by_hz is not None else (
            (t_num is not None and t_num < 0) or (db_num is not None and db_num < 0)
        )

        parts = iso.split("-")
        year = int(parts[0])
        month_tag = int(parts[1])
        day_num = int(parts[2])

        date_str = f"{day_num:02d}-{month_tag:02d}-{year}"

        rows_out.append({
            "date": iso,
            "date_iso": iso,
            "date_str": date_str,
            "year": year,
            "month_tag": month_tag,
            "tag": month_tag,
            "name": name_str,
            "amount": abs(t_num if t_num is not None else base),
            "debit": abs(db_num if db_num is not None else base),
            "__credit": bool(credit),
        })

    col_map = None  # current column mapping
    last_good_date = None
    heur = _detect_heuristic(matrix) or {}

    def row_has_any_amount(row):
        if col_map:
            probes = [col_map.get(k) for k in ("hCol", "zCol", "debitCol", "transCol", "hzCol") if col_map.get(k) is not None]
            for c in probes:
                if c < len(row):
                    n = numify(row[c])
                    if n is not None and abs(n) > 0:
                        return True
        for c in range(len(row)):
            n = numify(row[c])
            if n is not None and abs(n) > 0:
                return True
        return False

    def row_best_name(row):
        name = row[col_map["nameCol"]] if col_map and col_map.get("nameCol") is not None and col_map["nameCol"] < len(row) else None
        name = clean_html(name)
        name_str = str(name or "").strip()

        looks_like_numbers = bool(name_str and re.match(r"^[\d,.\s\-₪$€£]+$", name_str))
        looks_like_garbage = bool(name_str and re.match(r"^[x\s«»§=íï\-]+$", name_str, re.IGNORECASE))
        has_letters = bool(name_str and re.search(r"[A-Za-z\u0590-\u05FF]", name_str))
        needs_fallback = not name_str or re.match(r"^[:.\-_]+$", name_str) or looks_like_numbers or looks_like_garbage or not has_letters

        if needs_fallback and col_map:
            avoid = [col_map.get(k) for k in ("dateCol", "hCol", "zCol", "debitCol", "transCol", "hzCol")]
            name = _best_name_from_row(row, avoid)
        return str(name or "").strip()

    def looks_like_txn_row(row):
        if not row:
            return False
        has_date_any = any(parse_date_flex(v) for v in row) or last_good_date is not None
        has_name_any = any(
            re.search(r"[A-Za-z\u0590-\u05FF]", str(v or "").strip())
            and not _try_number(str(v or "").strip())
            for v in row
        )
        has_amt_any = any(numify(v) is not None and abs(numify(v)) > 0 for v in row)
        return has_name_any and has_amt_any and has_date_any

    for r_idx in range(len(matrix)):
        row = matrix[r_idx] if r_idx < len(matrix) else []
        if not row:
            continue

        # (1) Re-map on header row
        if _is_header_row(row):
            m = _map_from_header_row(matrix, r_idx)
            col_map = m
            last_good_date = None
            continue

        # Skip accidental header-looking rows
        if col_map and (
            col_map.get("dateCol") is not None and col_map["dateCol"] < len(row)
            and col_map.get("nameCol") is not None and col_map["nameCol"] < len(row)
            and col_map.get("debitCol") is not None and col_map["debitCol"] < len(row)
        ):
            dc = str(row[col_map["dateCol"]] or "")
            nc = str(row[col_map["nameCol"]] or "")
            dbc = str(row[col_map["debitCol"]] or "")
            if (re.search(r"[A-Za-z\u0590-\u05FF]", dc)
                and re.search(r"[A-Za-z\u0590-\u05FF]", nc)
                and re.search(r"[A-Za-z\u0590-\u05FF]", dbc)):
                continue

        # (2) Use heuristic if no header yet but row looks like data
        if col_map is None and looks_like_txn_row(row) and heur and (heur.get("dateCol") is not None or heur.get("nameCol") is not None):
            col_map = {
                "dateCol": heur.get("dateCol"),
                "nameCol": heur.get("nameCol"),
                "transCol": heur.get("transCol"),
                "debitCol": heur.get("debitCol"),
                "hCol": None,
                "zCol": None,
                "hzCol": None,
            }
            d_try = row[col_map["dateCol"]] if col_map.get("dateCol") is not None and col_map["dateCol"] < len(row) else None
            d0 = parse_date_flex(d_try)
            if d0:
                last_good_date = d0

        if col_map is None:
            continue

        # Skip totals / balance rows
        joined = " ".join(str(x or "") for x in row)
        if _TOTAL_RE.search(joined):
            continue

        # 1) Date from mapped column
        d_raw = row[col_map["dateCol"]] if col_map.get("dateCol") is not None and col_map["dateCol"] < len(row) else None
        d = parse_date_flex(d_raw)

        if not d and col_map.get("dateCol") is None:
            for c in range(min(3, len(row))):
                d = parse_date_flex(row[c])
                if d:
                    break

        # 2) Amounts
        credit_flag = None
        t, db = None, None

        if col_map.get("hCol") is not None or col_map.get("zCol") is not None:
            h = numify(row[col_map["hCol"]]) if col_map.get("hCol") is not None and col_map["hCol"] < len(row) else None
            z = numify(row[col_map["zCol"]]) if col_map.get("zCol") is not None and col_map["zCol"] < len(row) else None
            if h is not None and abs(h) > 0:
                t = h; db = h; credit_flag = False
            elif z is not None and abs(z) > 0:
                t = z; db = z; credit_flag = True

        if (t is None or db is None) and col_map.get("hzCol") is not None:
            hz_val = numify(row[col_map["hzCol"]]) if col_map["hzCol"] < len(row) else None
            if hz_val is not None and abs(hz_val) > 0:
                t = hz_val; db = hz_val
                credit_flag = hz_val > 0

        if (t is None or db is None) and col_map.get("debitCol") is not None:
            dval = numify(row[col_map["debitCol"]]) if col_map["debitCol"] < len(row) else None
            if dval is not None and abs(dval) > 0:
                db = dval
                if t is None:
                    t = dval

        if (t is None or db is None) and col_map.get("transCol") is not None:
            tval = numify(row[col_map["transCol"]]) if col_map["transCol"] < len(row) else None
            if tval is not None and abs(tval) > 0:
                t = tval
                if db is None:
                    db = tval

        if t is None and db is None:
            best_n, best_abs = None, 0
            known = [col_map.get(k) for k in ("hCol", "zCol", "debitCol", "transCol", "hzCol")]
            for c in range(len(row)):
                if c in known:
                    continue
                n = numify(row[c])
                if n is not None and abs(n) > best_abs:
                    best_n, best_abs = n, abs(n)
            if best_n is not None:
                t = best_n; db = best_n

        name = row_best_name(row)

        # 3) Carry-forward date
        if not d and name and (t is not None or db is not None) and row_has_any_amount(row) and last_good_date:
            d = last_good_date
        if d:
            last_good_date = d

        if not d or not name:
            continue

        push_row(d, name, t, db, credit_flag)

    return rows_out


def _try_number(s):
    """Return True if s is purely numeric (after stripping common symbols)."""
    cleaned = re.sub(r"[,\s₪$€£]", "", s)
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


# ---------------------------------------------------------------------------
# Workbook-level: read file bytes and dispatch by format
# ---------------------------------------------------------------------------

def _read_xlsx(file_bytes):
    """Parse XLSX bytes into list of (sheet_name, matrix) pairs."""
    if openpyxl is None:
        raise ImportError("openpyxl is required for XLSX files. Install: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        matrix = []
        for row in ws.iter_rows(values_only=True):
            matrix.append(list(row))
        sheets.append((name, matrix))
    wb.close()
    return sheets


def _read_xls(file_bytes, file_name):
    """Parse XLS bytes. Tries xlrd first, then BS4 for HTML-based XLS."""
    # First check if it's actually HTML disguised as .xls
    text_head = file_bytes[:500]
    is_html = False
    try:
        decoded = text_head.decode("utf-8", errors="ignore").lower()
        if "<html" in decoded or "<table" in decoded:
            is_html = True
    except Exception:
        pass

    if is_html:
        return _read_html_xls(file_bytes)

    if xlrd is None:
        raise ImportError("xlrd is required for binary XLS files. Install: pip install xlrd")
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sheets = []
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        matrix = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                        row.append(date(dt_tuple[0], dt_tuple[1], dt_tuple[2]))
                    except Exception:
                        row.append(cell.value)
                else:
                    row.append(cell.value)
            matrix.append(row)
        sheets.append((name, matrix))
    return sheets


def _read_html_xls(file_bytes):
    """Parse an HTML-based XLS file."""
    if BeautifulSoup is None:
        raise ImportError("beautifulsoup4 is required for HTML-based XLS. Install: pip install beautifulsoup4 lxml")

    # Try multiple encodings
    text = None
    for enc in ("utf-8", "windows-1255", "iso-8859-8", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        text = file_bytes.decode("utf-8", errors="replace")

    soup = BeautifulSoup(text, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        return [("Sheet1", [])]

    # Use the largest table
    best_table = max(tables, key=lambda t: len(t.find_all("tr")))
    matrix = []
    for tr in best_table.find_all("tr"):
        row = []
        for td in tr.find_all(["td", "th"]):
            cell_text = td.get_text(separator=" ", strip=True)
            row.append(cell_text)
        matrix.append(row)
    return [("Sheet1", matrix)]


def _read_csv(file_bytes):
    """Parse CSV bytes into a single sheet matrix."""
    # Try UTF-8 first, then fallbacks
    text = None
    for enc in ("utf-8-sig", "utf-8", "windows-1255", "iso-8859-8", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        text = file_bytes.decode("utf-8", errors="replace")

    # Detect dialect
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    matrix = [row for row in reader]
    return [("Sheet1", matrix)]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def extract_transactions(file_bytes, file_name, user_tag=None):
    """
    Main entry point: parse a file (bytes) into a list of transaction dicts.

    Args:
        file_bytes: raw file content as bytes
        file_name: original filename (for format detection)
        user_tag: optional month tag (1-12) for date validation

    Returns:
        list of dicts, each with keys:
            date, date_iso, date_str, year, month_tag, tag,
            name, amount, debit, __credit
    """
    ext = (file_name or "").rsplit(".", 1)[-1].lower()

    if ext == "xlsx":
        sheets = _read_xlsx(file_bytes)
    elif ext == "xls":
        sheets = _read_xls(file_bytes, file_name)
    elif ext == "csv":
        sheets = _read_csv(file_bytes)
    else:
        raise ValueError(f"Unsupported file format: .{ext}")

    all_rows = []
    for sheet_name, matrix in sheets:
        all_rows.extend(_extract_from_sheet(matrix, sheet_name, file_name, user_tag))

    logger.info(f"Parsed {len(all_rows)} rows from {file_name}")
    return all_rows
